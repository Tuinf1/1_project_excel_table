import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from pathlib import Path


# Загружаем исходные данные
df = pd.read_csv("data/orders.csv")

# Берём только уникальные значения external_id
unique_df = df.drop_duplicates(subset="external_id", keep="first")[["id", "external_id"]]

# Создаём матрицу статусов
status_pivot = (
    df.assign(flag=1)
      .pivot_table(index="external_id", columns="status", values="flag", aggfunc="max", fill_value=0)
      .reset_index()
)

# Если заказ доставлен — значит он прошёл все предыдущие стадии
delivered_mask = status_pivot["delivered"] == 1
cols_to_fill = ["created", "paid", "prod_started", "shipped"]
status_pivot.loc[delivered_mask, cols_to_fill] = 1


status_pivot["created"] = 1

df = status_pivot






# # Сохраняем в Excel
# out_path = "excel/unique_orders.xlsx"
# status_pivot.to_excel(out_path, sheet_name="Orders", index=False)

# # Автоширина колонок
# wb = load_workbook(out_path)
# ws = wb.active
# for col in ws.columns:
#     max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
#     ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
# wb.save(out_path)

# print(f"✅ Создан Excel-файл: {out_path}")
# print(f"📄 Строк: {len(status_pivot)} (уникальных external_id)")




s_created = df["created"].sum()
s_paid = df["paid"].sum()
s_prod_started = df["prod_started"].sum()
s_shipped = df["shipped"].sum()
s_delivered = df["delivered"].sum()



# Считаем отношения
# --- Считаем отношения и округляем ---
conv_pairs = [
    ("paid/created", round(s_paid / s_created, 4) if s_created else 0),
    ("prod_started/paid", round(s_prod_started / s_paid, 4) if s_paid else 0),
    ("shipped/prod_started", round(s_shipped / s_prod_started, 4) if s_prod_started else 0),
    ("delivered/shipped", round(s_delivered / s_shipped, 4) if s_shipped else 0),
    ("delivered/created", round(s_delivered / s_created, 4) if s_created else 0),
]


# --- Преобразуем в DataFrame ---
conv_df = pd.DataFrame(conv_pairs, columns=["stage", "ratio"])
# Загружаем Excel с конверсиями
df = conv_df




# === 5. Сохраняем в Excel ===
out_path = Path("excel/conversion_funnel.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
conv_df.to_excel(out_path, sheet_name="Funnel", index=False)

# === 6. Строим график ===
wb = load_workbook(out_path)
ws = wb.active

chart = BarChart()
# chart.title = "Conversion Funnel (по единицам)"
# chart.y_axis.title = "Конверсия"
# chart.x_axis.title = "Этап перехода"

data = Reference(ws, min_col=2, min_row=2, max_row=1 + len(conv_df))
cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(conv_df))
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)

from openpyxl.chart.label import DataLabelList

chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True
chart.dataLabels.numFmt = "0.0000"

ws.add_chart(chart, "E2")

# Автоширина
for col in ws.columns:
    max_len = max((len(str(c.value)) for c in col if c.value), default=0)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

wb.save(out_path)

print(f"✅ График создан и сохранён в: {out_path}")
print(conv_df)
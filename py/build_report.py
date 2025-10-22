import argparse
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList

# --- 1 Парсер аргументов ---
def parse_args():
    parser = argparse.ArgumentParser(description="Сбор Excel-отчёта из CSV и SQL")
    parser.add_argument("--days", type=int, default=90, help="Период в днях для выборки")
    parser.add_argument("--out", default="excel/Report.xlsx", help="Путь к итоговому Excel-файлу")
    parser.add_argument("--db", default=":memory:", help="Путь к SQLite базе или :memory:")
    return parser.parse_args()


# --- 2️⃣ Функция batch-вставки CSV ---
def insert_csv_to_db(conn, csv_path, table_name, batch_size=50000):
    df = pd.read_csv(csv_path)
    cols = list(df.columns)
    placeholders = ",".join(["?"] * len(cols))
    sql = f"INSERT INTO {table_name} ({','.join(cols)}) VALUES ({placeholders})"

    cur = conn.cursor()
    total = len(df)
    for start in range(0, total, batch_size):
        batch = df.iloc[start:start + batch_size].values.tolist()
        cur.executemany(sql, batch)
    conn.commit()
    # print(f"✅ {table_name}: вставлено {total:,} строк ({csv_path})")


# --- 3️⃣ Основная функция ---
def main():
    args = parse_args()
    print(f"→ days={args.days}, out={args.out}, db={args.db}")

    # --- SQLite ---
    conn = sqlite3.connect(args.db)
    cur = conn.cursor()
    cur.execute("PRAGMA journal_mode=WAL;")
    cur.execute("PRAGMA synchronous=NORMAL;")
    conn.commit()
    if args.db == ":memory:":
        # --- Создание таблиц ---
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS sellers (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY,
            external_id TEXT NOT NULL,
            date TEXT NOT NULL,
            channel TEXT NOT NULL,
            seller_id INTEGER,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            delivered_at TEXT,
            FOREIGN KEY(seller_id) REFERENCES sellers(id)
        );
        CREATE TABLE IF NOT EXISTS order_items (
            order_id INTEGER NOT NULL,
            sku TEXT NOT NULL,
            qty INTEGER NOT NULL,
            revenue REAL NOT NULL,
            cost REAL NOT NULL,
            FOREIGN KEY(order_id) REFERENCES orders(id)
        );
        """)

    # --- Очистка перед загрузкой ---
    cur.executescript("DELETE FROM order_items; DELETE FROM orders; DELETE FROM sellers;")
    conn.commit()
   

    # --- Загрузка CSV ---
    data_dir = Path("data")
    insert_csv_to_db(conn, data_dir / "sellers.csv", "sellers")
    insert_csv_to_db(conn, data_dir / "orders.csv", "orders")
    insert_csv_to_db(conn, data_dir / "order_items.csv", "order_items")

    # --- Индексы ---
    cur.executescript("""
    CREATE INDEX IF NOT EXISTS idx_orders_external_id ON orders(external_id);
    CREATE INDEX IF NOT EXISTS idx_orders_date ON orders(date);
    CREATE INDEX IF NOT EXISTS idx_orders_updated_at ON orders(updated_at);
    CREATE INDEX IF NOT EXISTS idx_orders_delivered_at ON orders(delivered_at);
    CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id);
    """)
    

    # --- Выполнение SQL-выгрузки ---
    sql_query = Path("sql/export.sql").read_text(encoding="utf-8")
    orders_df = pd.read_sql_query(sql_query, conn, params={"days": args.days})
    

    # --- 7️⃣ Summary ---
    summary_df = (
        orders_df.groupby(["channel", "seller"], as_index=False)
        .agg(
            revenue_sum=("revenue", "sum"),
            cost_sum=("cost", "sum"),
            margin_sum=("margin", "sum"),
            items_count=("sku", "count"),
            orders_count=("order_id", "nunique"),
        )
        .sort_values(["channel", "seller"])
    )



    # --- 8️⃣ Подсчёт статусов --
    orders_raw = pd.read_csv("data/orders.csv")
     # Матрица статусов
    status_pivot = (
        orders_raw.assign(flag=1)
        .pivot_table(index="external_id", columns="status", values="flag", aggfunc="max", fill_value=0)
        .reset_index()
    )

    if "delivered" in status_pivot.columns:
        delivered_mask = status_pivot["delivered"] == 1
        cols_to_fill = [c for c in ["created", "paid", "prod_started", "shipped"] if c in status_pivot.columns]
        status_pivot.loc[delivered_mask, cols_to_fill] = 1

    status_pivot["created"] = 1

    # Подсчёт по стадиям
    s_created = status_pivot["created"].sum()
    s_paid = status_pivot["paid"].sum()
    s_prod_started = status_pivot["prod_started"].sum()
    s_shipped = status_pivot["shipped"].sum()
    s_delivered = status_pivot["delivered"].sum()

    # Конверсии
    conv_pairs = [
        ("paid/created", round(s_paid / s_created, 4) if s_created else 0),
        ("prod_started/paid", round(s_prod_started / s_paid, 4) if s_paid else 0),
        ("shipped/prod_started", round(s_shipped / s_prod_started, 4) if s_prod_started else 0),
        ("delivered/shipped", round(s_delivered / s_shipped, 4) if s_shipped else 0),
        ("delivered/created", round(s_delivered / s_created, 4) if s_created else 0),
    ]
    conv_df = pd.DataFrame(conv_pairs, columns=["stage", "ratio"])

    

 


    
    # --- 10️⃣ Маржа по каналам ---
    margin_by_channel = (
        orders_df.groupby("channel", as_index=False)["margin"]
        .sum()
        .sort_values("margin", ascending=False)
    )

    # --- 11️⃣ Checks ---
    # print("🧩 Проверка данных...")
    bad_qty_df = orders_df.query("qty <= 0")[["order_id", "external_id", "sku", "qty", "status"]]
    bad_margin_df = orders_df.query("margin < 0")[["order_id", "external_id", "sku", "revenue", "cost", "margin", "status"]]
    missing_refs_df = orders_df[
        orders_df["seller"].isna() | orders_df["channel"].isna() | (orders_df["channel"] == "")
    ][["order_id", "external_id", "seller", "channel", "status"]]
    dup_external_df = (
        orders_df.groupby("external_id").size().reset_index(name="cnt").query("cnt > 1")
    )

    checks = {
        "qty_le_0": bad_qty_df,
        "margin_lt_0": bad_margin_df,
        "missing_refs": missing_refs_df,
        "duplicate_external_ids": dup_external_df
    }

    total_issues = sum(len(df) for df in checks.values())

    # --- 12️⃣ Формирование Excel ---
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # print(f"💾 Формируем Excel → {out_path}")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        orders_df.to_excel(writer, sheet_name="Orders", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        conv_df.to_excel(writer, sheet_name="Dashboard", index=False, startrow=0)
        margin_by_channel.to_excel(writer, sheet_name="Dashboard", index=False, startrow=len(conv_df) + 5)
        for name, df in checks.items():
            df.to_excel(writer, sheet_name="Checks", index=False, startrow=(len(df) + 3) * list(checks.keys()).index(name))

    wb = load_workbook(out_path)
    ws_dash = wb["Dashboard"]

   # --- График Conversion Funnel ---
    chart_conv = BarChart()
    chart_conv.title = "Conversion Funnel (по единицам)"
    chart_conv.y_axis.title = "Конверсия"
    chart_conv.x_axis.title = "Этап перехода"

    data = Reference(ws_dash, min_col=2, min_row=2, max_row=1 + len(conv_df))
    cats = Reference(ws_dash, min_col=1, min_row=2, max_row=1 + len(conv_df))
    chart_conv.add_data(data, titles_from_data=False)
    chart_conv.set_categories(cats)

    chart_conv.dataLabels = DataLabelList()
    chart_conv.dataLabels.showVal = True
    chart_conv.dataLabels.numFmt = "0.0000"

    ws_dash.add_chart(chart_conv, "E2")

    # График 2 — Margin by Channel
    start_margin = len(conv_df) + 5
    chart_margin = BarChart()
    chart_margin.title = "Margin by Channel"
    data2 = Reference(ws_dash, min_col=2, min_row=start_margin + 2, max_row=start_margin + 1 + len(margin_by_channel))
    cats2 = Reference(ws_dash, min_col=1, min_row=start_margin + 2, max_row=start_margin + 1 + len(margin_by_channel))
    chart_margin.add_data(data2, titles_from_data=False)
    chart_margin.set_categories(cats2)
    ws_dash.add_chart(chart_margin, "E20")

    # Автоширина
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    wb.save(out_path)
    print(total_issues)
    if total_issues > 0:
        print(f"❌ Найдены проблемы ({total_issues} строк). См. лист Checks.")
        exit(1)
    else:
        print(f"✅ Проверки пройдены. Отчёт готов: {out_path}")
        exit(0)


if __name__ == "__main__":
    main()
